import openpyxl

wb = openpyxl.load_workbook('six_flags_registration.xlsx')
sheet1 = wb.get_sheet_by_name('registrant_details_report (1)')
newsheet2 = wb.create_sheet("Sheet2")
sheet2 = newsheet2

maxc = sheet1.max_column
maxr = sheet1.max_row


# Student First Name
for i in range(2, maxr):
    first_name = sheet1.cell(row=i, column=12).value
    first_name = str(first_name)
    sheet2.cell(row=i, column=1).value = first_name.title()
    
# Student Last Name
for i in range(2, maxr):
    last_name = sheet1.cell(row=i, column=13).value
    last_name = str(last_name)
    sheet2.cell(row=i, column=2).value = last_name.title()

# grade
for i in range(2, maxr):
    grade = sheet1.cell(row=i, column=14).value
    sheet2.cell(row=i, column=3).value = grade

# Column 4 blank
    
# allergies
for i in range(2, maxr):
    allergy = sheet1.cell(row=i, column=15).value
    sheet2.cell(row=i, column=5).value = allergy

# insurance
for i in range(2, maxr):
    insur = sheet1.cell(row=i, column=10).value
    sheet2.cell(row=i, column=6).value = insur

# Column 7 blank    

# parent_full_name
for i in range(2, maxr):
    par_first_name = sheet1.cell(row=i, column=1).value
    par_first_name = str(par_first_name)
    par_last_name = sheet1.cell(row=i, column=2).value
    par_last_name = str(par_last_name)
    sheet2.cell(row=i, column=8).value = par_first_name.title() + ' ' + par_last_name.title()

# parent email
for i in range(2, maxr):
    par_email = sheet1.cell(row=i, column=3).value
    sheet2.cell(row=i, column=9).value = par_email

# phone
for i in range(2, maxr):
    par_phone = sheet1.cell(row=i, column=4).value
    sheet2.cell(row=i, column=10).value = par_phone

# full address
for i in range(2, maxr):
    street = sheet1.cell(row=i, column=5).value
    street = str(street)
    city = sheet1.cell(row=i, column=6).value
    city = str(city)
    state = sheet1.cell(row=i, column=7).value
    state = str(state)
    zip = sheet1.cell(row=i, column=8).value
    zip = str(zip)
    sheet2.cell(row=i, column=11).value = street + ', ' + city.title() + ', ' + state.title() + ', ' + zip

# Column 12 blank    

# emergency contact
for i in range(2, maxr):
    emerg = sheet1.cell(row=i, column=9).value
    sheet2.cell(row=i, column=13).value = emerg
    
wb.save('six_flags_registration.xlsx')
print('done')
